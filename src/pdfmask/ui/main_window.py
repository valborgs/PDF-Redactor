"""
메인 윈도우 UI 컴포넌트
"""

import os
import shutil
from datetime import datetime
from typing import Optional

import fitz
from PyQt6.QtWidgets import (
    QMainWindow,
    QWidget,
    QHBoxLayout,
    QTableWidget,
    QListWidget,
    QDockWidget,
    QToolBar,
    QTableWidgetItem,
    QFileDialog,
    QMessageBox,
)
from PyQt6.QtCore import Qt, QSize
from PyQt6.QtGui import QKeySequence, QAction
from PyQt6.QtWidgets import QStyle
from PyQt6.QtGui import QShortcut

from ..core.models import MaskEntry
from ..managers import (
    PdfDocumentManager,
    MaskDataManager,
    ProgressManager,
    LogManager,
)
from .pdf_view import ScrollablePdfView


class MainWindow(QMainWindow):
    """
    메인 윈도우 클래스
    
    PDF 마스킹 애플리케이션의 주 화면입니다.
    """

    def __init__(self) -> None:
        """초기화"""
        super().__init__()
        
        # PDF 문서 관리자
        self.pdf_manager = PdfDocumentManager()
        self.current_page_index: int = 0
        
        # 마스킹 데이터 관리자
        self.mask_data_manager = MaskDataManager()
        
        # 진행상황 관리자
        self.progress_manager = ProgressManager()
        
        # 로그 관리자
        self.log_manager = LogManager()
        self.log_manager.log_app_start()
        
        # 마스킹 데이터 저장
        self.masks: list[MaskEntry] = []
        
        # 폴더 내 PDF 파일 목록
        self.pdf_files: list[str] = []
        self.current_pdf_index: int = -1
        self.completed_files: list[str] = []  # 완료된 파일 리스트
        self.current_folder_path: str = ""  # 현재 작업 중인 폴더 경로
        
        # 백업 설정
        self.backup_enabled: bool = True  # 백업 활성화 여부 (기본값: 활성화)
        
        self.init_ui()
        self.setup_menu()
        self.setup_toolbar()
        self.setup_shortcuts()
        self.setup_statusbar()

    def init_ui(self) -> None:
        """UI 초기화"""
        self.setWindowTitle("PDF Mask - PDF 마스킹 프로그램")
        self.setGeometry(100, 100, 1400, 900)
        
        # Dock Widget 도킹 옵션 설정
        self.setDockOptions(
            QMainWindow.DockOption.AllowNestedDocks |
            QMainWindow.DockOption.AllowTabbedDocks |
            QMainWindow.DockOption.AnimatedDocks
        )

        # 중앙 위젯 설정
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # 메인 레이아웃 (수평)
        main_layout = QHBoxLayout()
        central_widget.setLayout(main_layout)

        # 가운데: PDF 페이지 표시 (스크롤 가능한 뷰)
        self.scrollable_pdf_view = ScrollablePdfView(self)
        self.pdf_view = self.scrollable_pdf_view.pdf_view
        
        # 시그널 연결
        self.pdf_view.maskCreated.connect(self.on_mask_created)
        
        main_layout.addWidget(self.scrollable_pdf_view)
        
        # 왼쪽: 마스킹 리스트 (Dock Widget)
        self.mask_list = QTableWidget()
        self.setup_mask_table()
        self.mask_dock_widget = QDockWidget("마스킹 리스트", self)
        self.mask_dock_widget.setWidget(self.mask_list)
        self.mask_dock_widget.setMinimumWidth(250)
        
        # Dock 기능 설정
        self.mask_dock_widget.setFeatures(
            QDockWidget.DockWidgetFeature.DockWidgetMovable |
            QDockWidget.DockWidgetFeature.DockWidgetFloatable |
            QDockWidget.DockWidgetFeature.DockWidgetClosable
        )
        
        # 도킹 영역 설정
        self.mask_dock_widget.setAllowedAreas(
            Qt.DockWidgetArea.LeftDockWidgetArea |
            Qt.DockWidgetArea.RightDockWidgetArea |
            Qt.DockWidgetArea.TopDockWidgetArea |
            Qt.DockWidgetArea.BottomDockWidgetArea
        )
        
        self.addDockWidget(Qt.DockWidgetArea.LeftDockWidgetArea, self.mask_dock_widget)

        # 오른쪽: PDF 파일 목록 (Dock Widget)
        self.pdf_file_list = QListWidget()
        self.pdf_file_list.itemDoubleClicked.connect(self.on_pdf_list_double_clicked)
        self.pdf_dock_widget = QDockWidget("PDF 파일 목록", self)
        self.pdf_dock_widget.setWidget(self.pdf_file_list)
        self.pdf_dock_widget.setMinimumWidth(200)
        
        # Dock 기능 설정
        self.pdf_dock_widget.setFeatures(
            QDockWidget.DockWidgetFeature.DockWidgetMovable |
            QDockWidget.DockWidgetFeature.DockWidgetFloatable |
            QDockWidget.DockWidgetFeature.DockWidgetClosable
        )
        
        # 도킹 영역 설정
        self.pdf_dock_widget.setAllowedAreas(
            Qt.DockWidgetArea.LeftDockWidgetArea |
            Qt.DockWidgetArea.RightDockWidgetArea |
            Qt.DockWidgetArea.TopDockWidgetArea |
            Qt.DockWidgetArea.BottomDockWidgetArea
        )
        
        self.addDockWidget(Qt.DockWidgetArea.RightDockWidgetArea, self.pdf_dock_widget)

    def setup_mask_table(self) -> None:
        """마스킹 리스트 테이블 설정"""
        # 컬럼 설정
        self.mask_list.setColumnCount(2)
        self.mask_list.setHorizontalHeaderLabels(["페이지", "메모"])
        
        # 크기 설정
        self.mask_list.setMinimumWidth(250)
        self.mask_list.setMaximumWidth(400)
        
        # 행 헤더 숨김
        self.mask_list.verticalHeader().setVisible(False)
        
        # 컬럼 크기 조절
        self.mask_list.setColumnWidth(0, 60)
        self.mask_list.horizontalHeader().setStretchLastSection(True)
        
        # 선택 모드: 행 단위 선택
        self.mask_list.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        
        # 편집 트리거: 더블클릭
        self.mask_list.setEditTriggers(
            QTableWidget.EditTrigger.DoubleClicked | 
            QTableWidget.EditTrigger.EditKeyPressed
        )
        
        # 시그널 연결
        self.mask_list.itemChanged.connect(self.on_mask_item_changed)

    def setup_menu(self) -> None:
        """메뉴바 설정"""
        menubar = self.menuBar()

        # 파일 메뉴
        file_menu = menubar.addMenu("파일(&F)")

        # PDF 열기 액션
        open_pdf_action = QAction("PDF 열기...", self)
        open_pdf_action.setShortcut(QKeySequence("Ctrl+O"))
        open_pdf_action.triggered.connect(self.open_pdf)
        file_menu.addAction(open_pdf_action)

        # 폴더 열기 액션
        open_folder_action = QAction("폴더 열기...", self)
        open_folder_action.setShortcut(QKeySequence("Ctrl+Shift+O"))
        open_folder_action.triggered.connect(self.open_folder)
        file_menu.addAction(open_folder_action)

        # 마스킹 내역 엑셀 저장 액션
        export_excel_action = QAction("마스킹 내역 엑셀 저장...", self)
        export_excel_action.triggered.connect(self.export_masks_to_excel)
        file_menu.addAction(export_excel_action)

        file_menu.addSeparator()

        # 종료 액션
        exit_action = QAction("종료", self)
        exit_action.setShortcut(QKeySequence("Ctrl+Q"))
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)
        
        # 설정 메뉴
        settings_menu = menubar.addMenu("설정(&S)")
        
        # 백업 활성화/비활성화 액션
        self.backup_toggle_action = QAction("PDF 백업 활성화", self)
        self.backup_toggle_action.setCheckable(True)
        self.backup_toggle_action.setChecked(self.backup_enabled)
        self.backup_toggle_action.triggered.connect(self.toggle_backup)
        settings_menu.addAction(self.backup_toggle_action)
        
        # 보기 메뉴
        view_menu = menubar.addMenu("보기(&V)")
        
        # 마스킹 리스트 토글 액션
        self.toggle_mask_list_action = self.mask_dock_widget.toggleViewAction()
        self.toggle_mask_list_action.setText("마스킹 리스트")
        view_menu.addAction(self.toggle_mask_list_action)
        
        # PDF 파일 목록 토글 액션
        self.toggle_pdf_list_action = self.pdf_dock_widget.toggleViewAction()
        self.toggle_pdf_list_action.setText("PDF 파일 목록")
        view_menu.addAction(self.toggle_pdf_list_action)

        # 도움말 메뉴
        help_menu = menubar.addMenu("도움말(&H)")

        # 단축키 안내
        shortcuts_action = QAction("단축키 안내", self)
        shortcuts_action.triggered.connect(self.show_shortcuts_help)
        help_menu.addAction(shortcuts_action)

        # 사용 방법
        usage_action = QAction("사용 방법", self)
        usage_action.triggered.connect(self.show_usage_help)
        help_menu.addAction(usage_action)

        # 정보
        about_action = QAction("정보", self)
        about_action.triggered.connect(self.show_about_dialog)
        help_menu.addAction(about_action)

    def setup_toolbar(self) -> None:
        """툴바 설정"""
        toolbar = QToolBar("메인 툴바")
        toolbar.setIconSize(QSize(24, 24))
        self.addToolBar(toolbar)

        # PDF 열기 버튼
        file_icon = self.style().standardIcon(QStyle.StandardPixmap.SP_FileIcon)
        open_pdf_action = QAction(file_icon, "PDF 열기", self)
        open_pdf_action.triggered.connect(self.open_pdf)
        toolbar.addAction(open_pdf_action)

        # 폴더 열기 버튼
        folder_icon = self.style().standardIcon(QStyle.StandardPixmap.SP_DirOpenIcon)
        open_folder_action = QAction(folder_icon, "폴더 열기", self)
        open_folder_action.triggered.connect(self.open_folder)
        toolbar.addAction(open_folder_action)

    def setup_shortcuts(self) -> None:
        """키보드 단축키 설정"""
        # 다음 페이지
        self.shortcut_next_page = QShortcut(QKeySequence(Qt.Key.Key_Right), self)
        self.shortcut_next_page.activated.connect(self.go_next_page)

        # 이전 페이지
        self.shortcut_prev_page = QShortcut(QKeySequence(Qt.Key.Key_Left), self)
        self.shortcut_prev_page.activated.connect(self.go_prev_page)

        # 다음 페이지 (PageDown)
        self.shortcut_page_down = QShortcut(QKeySequence(Qt.Key.Key_PageDown), self)
        self.shortcut_page_down.activated.connect(self.go_next_page)

        # 이전 페이지 (PageUp)
        self.shortcut_page_up = QShortcut(QKeySequence(Qt.Key.Key_PageUp), self)
        self.shortcut_page_up.activated.connect(self.go_prev_page)

        # 저장
        self.shortcut_save = QShortcut(QKeySequence("Ctrl+S"), self)
        self.shortcut_save.activated.connect(self.save_masks)
        
        # 삭제
        self.shortcut_delete = QShortcut(QKeySequence(Qt.Key.Key_Delete), self)
        self.shortcut_delete.activated.connect(self.delete_selected_mask)

        # 확대
        self.shortcut_zoom_in = QShortcut(QKeySequence(Qt.Key.Key_Plus), self)
        self.shortcut_zoom_in.activated.connect(self.zoom_in)

        # 축소
        self.shortcut_zoom_out = QShortcut(QKeySequence(Qt.Key.Key_Minus), self)
        self.shortcut_zoom_out.activated.connect(self.zoom_out)

    def setup_statusbar(self) -> None:
        """상태바 설정"""
        self.statusBar().showMessage("준비 (Ctrl + 드래그로 마스킹 영역 선택)")

    def update_page_view(self) -> None:
        """현재 페이지를 화면에 표시"""
        if self.pdf_manager.doc is None:
            return

        try:
            # 현재 페이지 객체 가져오기
            page = self.pdf_manager.doc[self.current_page_index]
            
            # 페이지 실제 크기
            page_width = page.rect.width
            page_height = page.rect.height
            
            # 줌 레벨 적용
            zoom = self.scrollable_pdf_view.zoom_level * 1.5
            
            # 페이지 렌더링
            pixmap = self.pdf_manager.get_page_pixmap(self.current_page_index, zoom)
            
            if pixmap is not None:
                # PdfPageView에 페이지 설정
                self.pdf_view.set_page(
                    self.current_page_index,
                    pixmap,
                    page_width,
                    page_height,
                    self.masks
                )
                
                # 상태바 업데이트
                total_pages = self.pdf_manager.get_page_count()
                zoom_percent = int(self.scrollable_pdf_view.zoom_level * 100)
                self.statusBar().showMessage(
                    f"페이지: {self.current_page_index + 1} / {total_pages} | "
                    f"확대: {zoom_percent}% | "
                    f"(Ctrl + 드래그: 마스킹, Ctrl + 휠: 확대/축소)"
                )
            else:
                self.statusBar().showMessage("오류: 페이지 렌더링 실패")
                
        except Exception as e:
            print(f"페이지 표시 오류: {str(e)}")
            self.statusBar().showMessage(f"오류: {str(e)}")

    def zoom_in(self) -> None:
        """PDF 확대"""
        self.scrollable_pdf_view.zoom_in()

    def zoom_out(self) -> None:
        """PDF 축소"""
        self.scrollable_pdf_view.zoom_out()

    def show_shortcuts_help(self) -> None:
        """단축키 안내 다이얼로그 표시"""
        shortcuts_text = (
            "<b>[파일 관련]</b><br>"
            "<b>Ctrl+O</b> : PDF 열기<br>"
            "<b>Ctrl+Shift+O</b> : 폴더 열기<br>"
            "<b>Ctrl+S</b> : 마스킹 저장<br><br>"
            "<b>[페이지 이동]</b><br>"
            "<b>→ / PageDown</b> : 다음 페이지<br>"
            "<b>← / PageUp</b> : 이전 페이지<br><br>"
            "<b>[마스킹]</b><br>"
            "<b>Ctrl+드래그</b> : 마스킹 영역 선택<br>"
            "<b>Del</b> : 선택된 마스킹 삭제<br><br>"
            "<b>[기타]</b><br>"
            "<b>Ctrl+Q</b> : 프로그램 종료"
        )
        QMessageBox.information(self, "단축키 안내", shortcuts_text)

    def show_usage_help(self) -> None:
        """사용 방법 안내 다이얼로그 표시"""
        usage_text = (
            "1. 상단 툴바 또는 파일 메뉴에서 PDF 또는 폴더를 엽니다.\n"
            "2. 좌측 '마스킹 리스트' 패널은 생성된 마스킹 정보를 보여줍니다.\n"
            "3. 중앙 PDF 화면에서 Ctrl 키를 누른 상태로 마우스를 드래그하여 마스킹 영역을 선택합니다.\n"
            "4. 우측 'PDF 파일 목록'에서 다른 PDF를 더블클릭하여 전환할 수 있습니다.\n"
            "5. 마스킹이 완료되면 Ctrl+S로 마스킹을 적용하고 저장합니다.\n"
        )
        QMessageBox.information(self, "사용 방법", usage_text)

    def show_about_dialog(self) -> None:
        """프로그램 정보 다이얼로그 표시"""
        about_text = (
            "PDF Mask v1.5\n\n"
            "PDF 문서의 민감한 정보를 마스킹하기 위한 PyQt6 기반 도구입니다.\n"
            "PyMuPDF를 사용하여 PDF를 렌더링하고, Redaction 기능으로 내용을 영구적으로 가립니다.\n\n"
            "제작: PDF Mask Development Team"
        )
        QMessageBox.information(self, "정보", about_text)

    def reload_current_pdf(self) -> None:
        """현재 PDF 파일을 다시 로드"""
        if self.pdf_manager.file_path is None:
            return
        
        try:
            current_path = self.pdf_manager.file_path
            current_page = self.current_page_index
            
            # PDF 재로드
            self.pdf_manager.load_pdf(current_path)
            
            # 현재 페이지로 이동
            page_count = self.pdf_manager.get_page_count()
            if current_page >= page_count:
                self.current_page_index = max(0, page_count - 1)
            else:
                self.current_page_index = current_page
            
            # 페이지 표시
            self.update_page_view()
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "오류",
                f"PDF 파일을 다시 로드할 수 없습니다.\n\n{str(e)}"
            )

    def clear_masks(self) -> None:
        """마스킹 데이터 및 테이블 초기화"""
        self.masks.clear()
        self.mask_list.setRowCount(0)

    def load_pdf_from_path(self, file_path: str) -> None:
        """지정된 경로의 PDF 파일 로드"""
        try:
            # PDF 로드
            self.pdf_manager.load_pdf(file_path)
            
            # 마스킹 데이터 초기화
            self.clear_masks()
            
            # 저장된 마스킹 데이터 로드 시도
            success, loaded_masks, msg = self.mask_data_manager.load_masks(file_path)
            if success and loaded_masks:
                self.masks = loaded_masks
                # 테이블에 마스킹 데이터 표시
                for mask in self.masks:
                    row = self.mask_list.rowCount()
                    self.mask_list.insertRow(row)
                    
                    # 페이지 컬럼
                    page_item = QTableWidgetItem(str(mask.page_index + 1))
                    page_item.setFlags(page_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    self.mask_list.setItem(row, 0, page_item)
                    
                    # 메모 컬럼
                    note_item = QTableWidgetItem(mask.note)
                    self.mask_list.setItem(row, 1, note_item)
                    
                    page_item.setData(Qt.ItemDataRole.UserRole, row)
                
                print(f"마스킹 데이터 로드: {msg}")
            
            # 첫 페이지로 이동
            self.current_page_index = 0
            
            # 줌 레벨 초기화
            self.scrollable_pdf_view.zoom_level = 1.0
            
            # 페이지 표시
            self.update_page_view()
            
            # 창 제목에 파일명 표시
            filename = os.path.basename(file_path)
            self.setWindowTitle(f"PDF Mask - {filename}")
            
            # 로그 기록
            self.log_manager.log_pdf_open(file_path)
            print(f"PDF 로드 성공: {file_path}")
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "오류",
                f"PDF 파일을 열 수 없습니다.\n\n{str(e)}"
            )
            print(f"PDF 로드 실패: {str(e)}")

    def load_pdf_from_list(self, index: int) -> None:
        """PDF 파일 목록에서 지정된 인덱스의 PDF 로드"""
        if index < 0 or index >= len(self.pdf_files):
            return
        
        self.current_pdf_index = index
        file_path = self.pdf_files[index]
        self.load_pdf_from_path(file_path)
        
        # 리스트에서 현재 파일 하이라이트
        self.pdf_file_list.setCurrentRow(index)

    def delete_selected_mask(self) -> None:
        """선택된 마스킹 항목 삭제"""
        selected_rows = self.mask_list.selectionModel().selectedRows()
        
        if not selected_rows:
            return
        
        # 역순으로 삭제
        rows_to_delete = sorted([index.row() for index in selected_rows], reverse=True)
        
        for row in rows_to_delete:
            if 0 <= row < len(self.masks):
                # 데이터에서 삭제
                del self.masks[row]
                # 테이블에서 삭제
                self.mask_list.removeRow(row)
        
        # 화면 갱신
        self.update_page_view()
        
        print(f"{len(rows_to_delete)}개의 마스킹 항목 삭제됨")

    def toggle_backup(self) -> None:
        """백업 활성화/비활성화 토글"""
        self.backup_enabled = self.backup_toggle_action.isChecked()
        status = "활성화" if self.backup_enabled else "비활성화"
        print(f"PDF 백업 {status}")

    def backup_current_pdf(self) -> tuple[bool, str]:
        """현재 PDF 파일을 백업 폴더에 복사"""
        if self.pdf_manager.file_path is None:
            return False, "백업할 PDF 파일이 없습니다."
        
        try:
            pdf_path = self.pdf_manager.file_path
            pdf_filename = os.path.basename(pdf_path)
            
            # 프로젝트 루트 경로 (src/pdfmask/ui/ -> src/ -> project_root/)
            current_file = os.path.abspath(__file__)
            ui_dir = os.path.dirname(current_file)        # ui/
            pdfmask_dir = os.path.dirname(ui_dir)         # pdfmask/
            src_dir = os.path.dirname(pdfmask_dir)        # src/
            project_root = os.path.dirname(src_dir)       # project_root/
            
            # 백업 폴더 생성
            today_str = datetime.now().strftime("%Y%m%d")
            backup_base = os.path.join(project_root, "backup")
            backup_dir = os.path.join(backup_base, today_str)
            
            os.makedirs(backup_dir, exist_ok=True)
            
            # 백업 파일 경로
            backup_path = os.path.join(backup_dir, pdf_filename)
            
            # 이미 백업된 파일이 있는지 확인
            if os.path.exists(backup_path):
                return True, f"이미 백업된 파일: {backup_path}"
            
            # 파일 복사
            shutil.copy2(pdf_path, backup_path)
            
            return True, backup_path
            
        except Exception as e:
            return False, f"백업 실패: {str(e)}"

    def export_masks_to_excel(self) -> None:
        """마스킹 작업 내역을 엑셀 파일로 저장"""
        if not self.masks:
            QMessageBox.information(self, "알림", "저장할 마스킹 내역이 없습니다.")
            return

        if self.pdf_manager.file_path is None:
            QMessageBox.warning(self, "경고", "열려 있는 PDF 파일 정보가 없습니다.")
            return

        # openpyxl 지연 로딩
        try:
            from openpyxl import Workbook, load_workbook
        except ImportError:
            QMessageBox.critical(
                self,
                "오류",
                "openpyxl 패키지가 설치되어 있지 않습니다.\n\n"
                "다음 명령으로 설치 후 다시 시도해주세요.\n"
                "pip install openpyxl"
            )
            return

        try:
            # 저장 경로 및 파일명 구성
            pdf_path = self.pdf_manager.file_path
            base_dir = os.path.dirname(pdf_path) or os.getcwd()
            today_str = datetime.now().strftime("%Y%m%d")
            filename = f"마스킹_작업내역_{today_str}.xlsx"
            save_path = os.path.join(base_dir, filename)

            # 워크북 생성 또는 로드
            if os.path.exists(save_path):
                wb = load_workbook(save_path)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "마스킹 내역"
                # 헤더 행
                ws.append(["작업일시", "PDF 파일명", "페이지", "마스킹 영역 좌표", "메모"])

            pdf_name = os.path.basename(pdf_path)

            # 마스크 데이터 추가
            for mask in self.masks:
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                page_number = mask.page_index + 1
                rect = mask.rect
                coords = f"({rect.x0:.2f}, {rect.y0:.2f}, {rect.x1:.2f}, {rect.y1:.2f})"
                note = mask.note or ""

                ws.append([timestamp, pdf_name, page_number, coords, note])

            # 파일 저장
            wb.save(save_path)

        except Exception as e:
            QMessageBox.critical(
                self,
                "저장 오류",
                f"엑셀 파일 저장 중 오류가 발생했습니다.\n\n{str(e)}"
            )

    def open_pdf(self) -> None:
        """PDF 파일 열기"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "PDF 파일 열기", "", "PDF Files (*.pdf);;All Files (*)"
        )
        
        if file_path:
            self.load_pdf_from_path(file_path)
            
            # 단일 파일 열기
            self.pdf_files = [file_path]
            self.current_pdf_index = 0
            
            # PDF 파일 리스트 업데이트
            self.pdf_file_list.clear()
            self.pdf_file_list.addItem(os.path.basename(file_path))
            self.pdf_file_list.setCurrentRow(0)

    def open_folder(self) -> None:
        """폴더 열기 및 PDF 파일 목록 로드"""
        folder_path = QFileDialog.getExistingDirectory(self, "폴더 선택")
        
        if not folder_path:
            return
        
        print(f"선택된 폴더: {folder_path}")
        
        # 폴더 내 PDF 파일 검색
        pdf_files = []
        try:
            for filename in os.listdir(folder_path):
                if filename.lower().endswith('.pdf'):
                    full_path = os.path.join(folder_path, filename)
                    pdf_files.append(full_path)
        except Exception as e:
            QMessageBox.critical(
                self,
                "오류",
                f"폴더를 읽을 수 없습니다.\n\n{str(e)}"
            )
            return
        
        # PDF 파일이 없는 경우
        if not pdf_files:
            QMessageBox.information(
                self,
                "알림",
                "선택한 폴더에 PDF 파일이 없습니다."
            )
            return
        
        # 파일명으로 정렬
        pdf_files.sort()
        
        # PDF 파일 목록 저장
        self.pdf_files = pdf_files
        self.current_folder_path = folder_path
        self.completed_files = []
        
        # 이전 진행상황 확인
        success, progress_data, msg = self.progress_manager.load_progress()
        if success and progress_data and progress_data.get('folder_path') == folder_path:
            # 이전에 작업하던 폴더와 동일한 경우
            reply = QMessageBox.question(
                self,
                "진행상황 복구",
                f"이전 작업 진행상황이 있습니다.\n\n"
                f"완료: {progress_data.get('completed_count', 0)}/{progress_data.get('total_files', 0)}\n"
                f"마지막 작업: {progress_data.get('last_updated', '')}\n\n"
                f"이어서 작업하시겠습니까?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.Yes
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                # 진행상황 복구
                self.completed_files = progress_data.get('completed_files', [])
                start_index = progress_data.get('current_index', 0)
                
                print(f"진행상황 복구: {len(self.completed_files)}개 파일 완료")
        
        # 리스트 위젯 업데이트
        self.pdf_file_list.clear()
        for file_path in pdf_files:
            filename = os.path.basename(file_path)
            item_text = filename
            # 완료된 파일 표시
            if filename in self.completed_files:
                item_text = f"✓ {filename}"
            self.pdf_file_list.addItem(item_text)
        
        # 로그 기록
        self.log_manager.log_folder_open(folder_path, len(pdf_files))
        print(f"총 {len(pdf_files)}개의 PDF 파일 발견")
        
        # 첫 번째 PDF 또는 복구된 위치에서 시작
        if pdf_files:
            if success and progress_data and progress_data.get('folder_path') == folder_path:
                start_index = progress_data.get('current_index', 0)
                if 0 <= start_index < len(pdf_files):
                    self.load_pdf_from_list(start_index)
                else:
                    self.load_pdf_from_list(0)
            else:
                self.load_pdf_from_list(0)

    def on_pdf_list_double_clicked(self, item) -> None:
        """PDF 파일 리스트에서 더블클릭 이벤트"""
        row = self.pdf_file_list.row(item)
        self.load_pdf_from_list(row)

    def go_next_page(self) -> None:
        """다음 페이지로 이동"""
        if self.pdf_manager.doc is None:
            return
        
        page_count = self.pdf_manager.get_page_count()
        if self.current_page_index < page_count - 1:
            self.current_page_index += 1
            self.update_page_view()
            print(f"다음 페이지: {self.current_page_index + 1}")

    def go_prev_page(self) -> None:
        """이전 페이지로 이동"""
        if self.pdf_manager.doc is None:
            return
        
        if self.current_page_index > 0:
            self.current_page_index -= 1
            self.update_page_view()
            print(f"이전 페이지: {self.current_page_index + 1}")

    def save_masks(self) -> None:
        """마스킹 정보 저장"""
        if self.pdf_manager.doc is None:
            return
        
        # 마스크가 있는 경우
        if len(self.masks) > 0:
            reply = QMessageBox.question(
                self,
                "마스킹 저장",
                "현재 PDF 파일의 마스킹 작업 내용을 저장하시겠습니까?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.Yes
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                try:
                    # 백업 수행
                    if self.backup_enabled:
                        backup_success, backup_msg = self.backup_current_pdf()
                        
                        if not backup_success:
                            retry_reply = QMessageBox.warning(
                                self,
                                "백업 실패",
                                f"{backup_msg}\n\n백업 없이 계속 진행하시겠습니까?",
                                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                                QMessageBox.StandardButton.No
                            )
                            
                            if retry_reply == QMessageBox.StandardButton.No:
                                return
                        else:
                            print(f"백업 완료: {backup_msg}")
                    
                    # 마스킹 데이터 JSON 저장
                    json_success, json_msg = self.mask_data_manager.save_masks(
                        self.pdf_manager.file_path,
                        self.masks
                    )
                    if json_success:
                        print(f"마스킹 데이터 JSON 저장: {json_msg}")
                    
                    # 마스킹 적용 및 저장
                    self.pdf_manager.apply_masks_and_save(self.masks)
                    
                    # 마스킹 내역 엑셀 저장
                    self.export_masks_to_excel()
                    
                    # 로그 기록
                    self.log_manager.log_mask_save(
                        self.pdf_manager.file_path,
                        self.masks
                    )
                    
                    # 성공 메시지
                    QMessageBox.information(
                        self,
                        "저장 완료",
                        f"총 {len(self.masks)}개의 마스킹이 적용되어 저장되었습니다."
                    )
                    
                    # 마스킹 데이터 초기화
                    self.clear_masks()
                    
                    # 완료 파일 목록에 추가
                    if self.pdf_manager.file_path:
                        filename = os.path.basename(self.pdf_manager.file_path)
                        if filename not in self.completed_files:
                            self.completed_files.append(filename)
                    
                    # 진행상황 저장
                    if self.current_folder_path and self.pdf_files:
                        self.progress_manager.save_progress(
                            self.current_folder_path,
                            self.pdf_files,
                            self.completed_files,
                            self.current_pdf_index
                        )
                    
                    # 다음 파일로 이동할지 확인
                    self.move_to_next_pdf_if_available()
                    
                    print("마스킹 저장 완료")
                    
                except Exception as e:
                    QMessageBox.critical(
                        self,
                        "저장 오류",
                        f"저장 중 오류가 발생했습니다.\n\n{str(e)}"
                    )
                    print(f"저장 실패: {str(e)}")
        
        # 마스크가 없는 경우
        else:
            reply = QMessageBox.question(
                self,
                "다음으로 이동",
                "마스킹 작업 없이 다음으로 넘어가시겠습니까?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.Yes
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                self.move_to_next_pdf_if_available()

    def move_to_next_pdf_if_available(self) -> None:
        """폴더의 다음 PDF 파일로 이동"""
        if self.current_pdf_index >= 0 and self.current_pdf_index < len(self.pdf_files) - 1:
            next_index = self.current_pdf_index + 1
            
            reply = QMessageBox.question(
                self,
                "다음 파일",
                f"다음 PDF 파일을 열겠습니까?\n\n{os.path.basename(self.pdf_files[next_index])}",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.Yes
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                self.load_pdf_from_list(next_index)
        else:
            # 마지막 파일인 경우
            if self.current_pdf_index == len(self.pdf_files) - 1:
                # 진행상황 삭제
                self.progress_manager.clear_progress()
                
                QMessageBox.information(
                    self,
                    "완료",
                    "폴더의 모든 PDF 파일 작업이 완료되었습니다."
                )

    def on_mask_created(self, page_index: int, rect: fitz.Rect) -> None:
        """마스킹 영역이 생성되었을 때 호출되는 슬롯"""
        # 새로운 MaskEntry 생성
        mask_entry = MaskEntry(page_index=page_index, rect=rect, note="")
        self.masks.append(mask_entry)
        
        print(f"Mask created on page {page_index + 1}: {rect}")
        
        # 테이블에 행 추가
        row = self.mask_list.rowCount()
        self.mask_list.insertRow(row)
        
        # 페이지 컬럼
        page_item = QTableWidgetItem(str(page_index + 1))
        page_item.setFlags(page_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
        self.mask_list.setItem(row, 0, page_item)
        
        # 메모 컬럼
        note_item = QTableWidgetItem("")
        self.mask_list.setItem(row, 1, note_item)
        
        # UserRole에 mask 인덱스 저장
        page_item.setData(Qt.ItemDataRole.UserRole, row)

    def on_mask_item_changed(self, item: QTableWidgetItem) -> None:
        """마스킹 리스트 아이템이 변경되었을 때 호출되는 슬롯"""
        row = item.row()
        col = item.column()
        
        # 메모 컬럼만 처리
        if col == 1:
            if 0 <= row < len(self.masks):
                new_note = item.text()
                self.masks[row].note = new_note
                print(f"Mask [{row}] note updated: '{new_note}'")

    def closeEvent(self, event) -> None:
        """윈도우 종료 이벤트"""
        # 로그 기록
        self.log_manager.log_app_end()
        
        # PDF 문서 닫기
        self.pdf_manager.close()
        event.accept()

