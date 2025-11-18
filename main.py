"""
PDF Mask - PDF 마스킹 프로그램
PyQt6 기반 GUI 애플리케이션
"""

import sys
import os
import shutil
import json
import hashlib
import logging
from typing import Optional
from dataclasses import dataclass
from datetime import datetime
import fitz  # PyMuPDF
from PyQt6.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QHBoxLayout,
    QVBoxLayout,
    QTableWidget,
    QLabel,
    QListWidget,
    QDockWidget,
    QToolBar,
    QTableWidgetItem,
    QFileDialog,
    QMessageBox,
    QScrollArea,
    QDialog,
    QLineEdit,
    QPushButton,
    QProgressDialog,
)
from PyQt6.QtCore import Qt, QSize, QPoint, QRect, pyqtSignal
from PyQt6.QtGui import QKeySequence, QAction, QImage, QPixmap, QPainter, QColor, QPen, QBrush, QIcon
from PyQt6.QtWidgets import QStyle


@dataclass
class MaskEntry:
    """마스킹 정보 데이터 클래스"""
    page_index: int
    rect: fitz.Rect
    note: str = ""


class LicenseManager:
    """라이선스 관리 클래스"""
    
    def __init__(self) -> None:
        # 프로젝트 루트 경로
        self.project_root = os.path.dirname(os.path.abspath(__file__))
        self.license_file = os.path.join(self.project_root, ".license")
    
    def is_licensed(self) -> bool:
        """라이선스가 유효한지 확인"""
        if not os.path.exists(self.license_file):
            return False
        
        try:
            with open(self.license_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return data.get('activated', False)
        except Exception:
            return False
    
    def validate_serial(self, serial: str) -> tuple[bool, str]:
        """
        시리얼 번호 검증 (간단한 로컬 검증)
        
        Args:
            serial: 입력된 시리얼 번호
            
        Returns:
            tuple[bool, str]: (검증 성공 여부, 메시지)
        """
        # 시리얼 번호 형식 검증 (예: XXXX-XXXX-XXXX-XXXX)
        serial = serial.strip().upper()
        
        if len(serial) == 0:
            return False, "시리얼 번호를 입력해주세요."
        
        # 간단한 형식 검증
        parts = serial.split('-')
        if len(parts) != 4:
            return False, "올바른 형식이 아닙니다. (XXXX-XXXX-XXXX-XXXX)"
        
        for part in parts:
            if len(part) != 4:
                return False, "올바른 형식이 아닙니다. (XXXX-XXXX-XXXX-XXXX)"
        
        return True, "형식 검증 성공"
    
    def activate_license(self, serial: str) -> tuple[bool, str]:
        """
        라이선스 활성화 (서버 인증 시뮬레이션)
        
        실제 구현에서는 여기서 서버 API를 호출해야 합니다.
        현재는 로컬에서 간단한 검증만 수행합니다.
        
        Args:
            serial: 시리얼 번호
            
        Returns:
            tuple[bool, str]: (활성화 성공 여부, 메시지)
        """
        # 1. 형식 검증
        valid, msg = self.validate_serial(serial)
        if not valid:
            return False, msg
        
        # 2. 서버 인증 시뮬레이션
        # 실제 구현: requests.post('https://license-server.com/activate', ...)
        
        # 간단한 해시 기반 검증 (예시)
        serial_hash = hashlib.sha256(serial.encode()).hexdigest()
        
        # 테스트용 시리얼: TEST-1234-5678-ABCD
        valid_serials = [
            "TEST-1234-5678-ABCD",
            "DEMO-0000-0000-0001",
        ]
        
        if serial not in valid_serials:
            return False, "유효하지 않은 시리얼 번호입니다."
        
        # 3. 라이선스 파일 저장
        try:
            license_data = {
                'activated': True,
                'serial': serial,
                'serial_hash': serial_hash,
                'activated_at': datetime.now().isoformat(),
            }
            
            with open(self.license_file, 'w', encoding='utf-8') as f:
                json.dump(license_data, f, indent=2, ensure_ascii=False)
            
            return True, "라이선스 활성화 완료"
            
        except Exception as e:
            return False, f"라이선스 저장 실패: {str(e)}"
    
    def deactivate_license(self) -> None:
        """라이선스 비활성화"""
        if os.path.exists(self.license_file):
            try:
                os.remove(self.license_file)
            except Exception:
                pass


class SerialInputDialog(QDialog):
    """시리얼 번호 입력 다이얼로그"""
    
    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.license_manager = LicenseManager()
        self.init_ui()
    
    def init_ui(self) -> None:
        """UI 초기화"""
        self.setWindowTitle("라이선스 인증")
        self.setModal(True)
        self.setFixedSize(400, 200)
        
        layout = QVBoxLayout()
        
        # 안내 문구
        info_label = QLabel(
            "프로그램을 사용하려면 시리얼 번호를 입력해주세요.\n\n"
            "형식: XXXX-XXXX-XXXX-XXXX"
        )
        info_label.setWordWrap(True)
        layout.addWidget(info_label)
        
        # 시리얼 번호 입력
        self.serial_input = QLineEdit()
        self.serial_input.setPlaceholderText("시리얼 번호 입력")
        self.serial_input.setMaxLength(19)  # XXXX-XXXX-XXXX-XXXX
        layout.addWidget(self.serial_input)
        
        # 버튼 레이아웃
        button_layout = QHBoxLayout()
        
        # 인증 버튼
        self.activate_button = QPushButton("인증")
        self.activate_button.clicked.connect(self.activate)
        button_layout.addWidget(self.activate_button)
        
        # 취소 버튼
        cancel_button = QPushButton("취소")
        cancel_button.clicked.connect(self.reject)
        button_layout.addWidget(cancel_button)
        
        layout.addLayout(button_layout)
        
        # 상태 메시지
        self.status_label = QLabel("")
        self.status_label.setStyleSheet("color: red;")
        layout.addWidget(self.status_label)
        
        self.setLayout(layout)
    
    def activate(self) -> None:
        """인증 버튼 클릭"""
        serial = self.serial_input.text().strip()
        
        if not serial:
            self.status_label.setText("시리얼 번호를 입력해주세요.")
            return
        
        # 진행 다이얼로그 표시
        progress = QProgressDialog("라이선스 서버에 연결 중...", None, 0, 0, self)
        progress.setWindowTitle("인증 중")
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.show()
        QApplication.processEvents()
        
        # 라이선스 활성화 시도
        success, message = self.license_manager.activate_license(serial)
        
        progress.close()
        
        if success:
            QMessageBox.information(
                self,
                "인증 성공",
                "라이선스가 성공적으로 활성화되었습니다."
            )
            self.accept()
        else:
            self.status_label.setText(message)
            QMessageBox.warning(
                self,
                "인증 실패",
                message
            )


class PdfPageView(QWidget):
    """PDF 페이지를 표시하고 마스킹 영역을 선택할 수 있는 커스텀 위젯"""
    
    # 시그널: (page_index, fitz.Rect)
    maskCreated = pyqtSignal(int, object)
    
    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        
        # 페이지 표시 관련
        self._pixmap: Optional[QPixmap] = None
        self._page_index: int = 0
        self._page_width: float = 0.0
        self._page_height: float = 0.0
        
        # 마스킹 드래그 관련
        self._start_pos: Optional[QPoint] = None
        self._current_rect: Optional[QRect] = None
        self._ctrl_pressed_during_drag: bool = False
        
        # 저장된 마스킹 영역들 (화면 좌표)
        self._saved_masks: list[QRect] = []
        
        # 줌 레벨
        self._zoom_level: float = 1.0  # 100%
        
        # 배경 스타일 (PDF 영역은 흰색)
        self.setStyleSheet("background-color: #ffffff;")
        self.setMinimumSize(400, 300)
        
    def set_page(
        self, 
        page_index: int, 
        pixmap: QPixmap, 
        page_width: float, 
        page_height: float,
        masks: list = None
    ) -> None:
        """페이지 정보 설정"""
        self._page_index = page_index
        self._pixmap = pixmap
        self._page_width = page_width
        self._page_height = page_height
        
        # 저장된 마스킹 영역 표시를 위해 변환
        self._saved_masks = []
        if masks and pixmap:
            for mask in masks:
                if mask.page_index == page_index:
                    # PDF 좌표를 화면 좌표로 변환
                    screen_rect = self._convert_to_screen_rect(mask.rect)
                    if screen_rect:
                        self._saved_masks.append(screen_rect)
        
        # 위젯 크기를 pixmap 크기에 맞춤
        if pixmap is not None:
            self.setFixedSize(pixmap.size())
        
        self.update()
    
    def set_zoom_level(self, zoom: float) -> None:
        """줌 레벨 설정"""
        self._zoom_level = zoom
    
    def paintEvent(self, event) -> None:
        """페이지 렌더링"""
        painter = QPainter(self)
        
        # Pixmap 그리기
        if self._pixmap is not None:
            painter.drawPixmap(0, 0, self._pixmap)
        
        # 저장된 마스킹 영역 그리기 (반투명 빨간색)
        if self._saved_masks:
            brush = QBrush(QColor(255, 0, 0, 60))
            painter.setBrush(brush)
            pen = QPen(QColor(255, 0, 0), 2)
            painter.setPen(pen)
            
            for rect in self._saved_masks:
                painter.drawRect(rect)
        
        # 드래그 중인 사각형 그리기 (반투명 파란색)
        if self._current_rect is not None:
            brush = QBrush(QColor(0, 120, 255, 80))
            painter.setBrush(brush)
            pen = QPen(QColor(0, 120, 255), 2)
            painter.setPen(pen)
            
            painter.drawRect(self._current_rect)
        
        painter.end()
    
    def mousePressEvent(self, event) -> None:
        """마우스 클릭 시작"""
        # Ctrl + 좌클릭인 경우만 드래그 시작
        if (event.button() == Qt.MouseButton.LeftButton and 
            event.modifiers() & Qt.KeyboardModifier.ControlModifier):
            
            self._start_pos = event.pos()
            self._current_rect = None
            self._ctrl_pressed_during_drag = True
    
    def mouseMoveEvent(self, event) -> None:
        """마우스 드래그 중"""
        if self._ctrl_pressed_during_drag and self._start_pos is not None:
            # 시작점과 현재점으로 사각형 생성
            current_pos = event.pos()
            
            # QRect 생성 (좌상단, 우하단을 자동으로 정규화)
            self._current_rect = QRect(self._start_pos, current_pos).normalized()
            
            # 화면 업데이트
            self.update()
    
    def mouseReleaseEvent(self, event) -> None:
        """마우스 클릭 종료"""
        if (self._ctrl_pressed_during_drag and 
            self._start_pos is not None and 
            self._current_rect is not None):
            
            # 화면 좌표를 PDF 페이지 좌표로 변환
            pdf_rect = self._convert_to_pdf_rect(self._current_rect)
            
            if pdf_rect is not None:
                # 시그널 발생
                self.maskCreated.emit(self._page_index, pdf_rect)
                
                # 저장된 마스킹 목록에 추가 (화면에 즉시 표시)
                self._saved_masks.append(self._current_rect)
            
            # 상태 초기화
            self._start_pos = None
            self._current_rect = None
            self._ctrl_pressed_during_drag = False
            
            # 화면 업데이트
            self.update()

    def wheelEvent(self, event) -> None:
        """마우스 휠 이벤트 (Ctrl + 휠로 줌)"""
        if event.modifiers() & Qt.KeyboardModifier.ControlModifier:
            # Ctrl이 눌린 상태에서는 확대/축소만 처리
            delta = event.angleDelta().y()
            parent = self.parent()
            if delta > 0:
                if parent is not None and hasattr(parent, "zoom_in"):
                    parent.zoom_in()
            else:
                if parent is not None and hasattr(parent, "zoom_out"):
                    parent.zoom_out()
            event.accept()
        else:
            # Ctrl이 아닐 때는 기본 스크롤 동작 유지
            super().wheelEvent(event)
    
    def _convert_to_pdf_rect(self, screen_rect: QRect) -> Optional[fitz.Rect]:
        """화면 좌표를 PDF 페이지 좌표로 변환"""
        if self._pixmap is None or self._page_width == 0 or self._page_height == 0:
            return None
        
        # 스케일 비율 계산
        scale_x = self._page_width / self._pixmap.width()
        scale_y = self._page_height / self._pixmap.height()
        
        # 화면 좌표를 PDF 좌표로 변환
        x0 = screen_rect.left() * scale_x
        y0 = screen_rect.top() * scale_y
        x1 = screen_rect.right() * scale_x
        y1 = screen_rect.bottom() * scale_y
        
        # fitz.Rect 생성 (x0, y0, x1, y1)
        return fitz.Rect(x0, y0, x1, y1)
    
    def _convert_to_screen_rect(self, pdf_rect: fitz.Rect) -> Optional[QRect]:
        """PDF 좌표를 화면 좌표로 변환"""
        if self._pixmap is None or self._page_width == 0 or self._page_height == 0:
            return None
        
        # 스케일 비율 계산
        scale_x = self._pixmap.width() / self._page_width
        scale_y = self._pixmap.height() / self._page_height
        
        # PDF 좌표를 화면 좌표로 변환
        x0 = int(pdf_rect.x0 * scale_x)
        y0 = int(pdf_rect.y0 * scale_y)
        x1 = int(pdf_rect.x1 * scale_x)
        y1 = int(pdf_rect.y1 * scale_y)
        
        return QRect(QPoint(x0, y0), QPoint(x1, y1))


class ScrollablePdfView(QScrollArea):
    """스크롤 가능한 PDF 뷰 컨테이너"""
    
    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        
        # PDF 뷰 위젯
        self.pdf_view = PdfPageView(self)
        self.setWidget(self.pdf_view)
        
        # 스크롤 설정
        self.setWidgetResizable(False)  # 위젯 크기 자동 조절 끄기
        self.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        
        # 크기 정책 설정
        # 가로: 최소 400px, 세로: 부모 창 크기에 맞춤
        self.setMinimumWidth(400)
        
        # PDF를 가운데 정렬
        self.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        # 배경색 설정 (회색)
        self.setStyleSheet("QScrollArea { background-color: #808080; }")
        
        # 줌 레벨
        self.zoom_level: float = 1.0  # 100%
        self.min_zoom: float = 0.5  # 최소 (높이에 맞추기용)
        self.max_zoom: float = 2.0  # 최대 200%
        
    def zoom_in(self) -> None:
        """줌 인 (10% 증가)"""
        if self.zoom_level < self.max_zoom:
            self.zoom_level = min(self.zoom_level + 0.1, self.max_zoom)
            if hasattr(self.parent(), 'update_page_view'):
                self.parent().update_page_view()
    
    def zoom_out(self) -> None:
        """줌 아웃 (10% 감소)"""
        if self.zoom_level > self.min_zoom:
            self.zoom_level = max(self.zoom_level - 0.1, self.min_zoom)
            if hasattr(self.parent(), 'update_page_view'):
                self.parent().update_page_view()

    def wheelEvent(self, event) -> None:
        """마우스 휠 이벤트 (Ctrl + 휠로 줌)"""
        if event.modifiers() & Qt.KeyboardModifier.ControlModifier:
            delta = event.angleDelta().y()
            if delta > 0:
                self.zoom_in()
            else:
                self.zoom_out()
            event.accept()
        else:
            # 기본 스크롤 동작 유지
            super().wheelEvent(event)


class LogManager:
    """로그 관리 클래스"""
    
    def __init__(self) -> None:
        # 프로젝트 루트 경로
        self.project_root = os.path.dirname(os.path.abspath(__file__))
        self.logs_dir = os.path.join(self.project_root, "logs")
        
        # 로그 폴더 생성
        os.makedirs(self.logs_dir, exist_ok=True)
        
        # 일자별 로그 파일 설정
        self.setup_logger()
    
    def setup_logger(self) -> None:
        """로거 설정 (일자별 로그 파일)"""
        today_str = datetime.now().strftime("%Y%m%d")
        log_filename = f"pdfmask_{today_str}.log"
        log_filepath = os.path.join(self.logs_dir, log_filename)
        
        # 기존 핸들러 제거
        logger = logging.getLogger('PDFMask')
        logger.handlers.clear()
        
        # 로거 레벨 설정
        logger.setLevel(logging.INFO)
        
        # 파일 핸들러
        file_handler = logging.FileHandler(log_filepath, encoding='utf-8')
        file_handler.setLevel(logging.INFO)
        
        # 포맷 설정
        formatter = logging.Formatter(
            '%(asctime)s - %(levelname)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        file_handler.setFormatter(formatter)
        
        logger.addHandler(file_handler)
        
        self.logger = logger
    
    def info(self, message: str) -> None:
        """정보 로그"""
        self.logger.info(message)
    
    def warning(self, message: str) -> None:
        """경고 로그"""
        self.logger.warning(message)
    
    def error(self, message: str) -> None:
        """에러 로그"""
        self.logger.error(message)
    
    def log_app_start(self) -> None:
        """프로그램 시작 로그"""
        self.info("=" * 60)
        self.info("PDF Mask Application Started")
        self.info("=" * 60)
    
    def log_app_end(self) -> None:
        """프로그램 종료 로그"""
        self.info("=" * 60)
        self.info("PDF Mask Application Closed")
        self.info("=" * 60)
    
    def log_license_check(self, success: bool, message: str) -> None:
        """라이선스 인증 로그"""
        if success:
            self.info(f"License Check: SUCCESS - {message}")
        else:
            self.warning(f"License Check: FAILED - {message}")
    
    def log_pdf_open(self, file_path: str) -> None:
        """PDF 파일 열기 로그"""
        self.info(f"PDF Opened: {file_path}")
    
    def log_folder_open(self, folder_path: str, file_count: int) -> None:
        """폴더 열기 로그"""
        self.info(f"Folder Opened: {folder_path} ({file_count} PDF files)")
    
    def log_mask_save(self, file_path: str, masks: list) -> None:
        """마스킹 저장 로그"""
        self.info(f"Mask Saved: {file_path} ({len(masks)} masks)")
        
        # 각 마스킹 영역 상세 정보 기록
        for i, mask in enumerate(masks, 1):
            rect = mask.rect
            note = mask.note if mask.note else "(no note)"
            self.info(
                f"  Mask #{i}: Page {mask.page_index + 1}, "
                f"Rect({rect.x0:.2f}, {rect.y0:.2f}, {rect.x1:.2f}, {rect.y1:.2f}), "
                f"Note: {note}"
            )
    
    def log_error(self, operation: str, error_message: str) -> None:
        """에러 로그"""
        self.error(f"Error in {operation}: {error_message}")


class ProgressManager:
    """작업 진행상황 관리 클래스"""
    
    def __init__(self) -> None:
        # 프로젝트 루트 경로
        self.project_root = os.path.dirname(os.path.abspath(__file__))
        self.progress_file = os.path.join(self.project_root, "progress.json")
    
    def save_progress(self, folder_path: str, pdf_files: list[str], completed_files: list[str], current_index: int) -> tuple[bool, str]:
        """
        작업 진행상황 저장
        
        Args:
            folder_path: 작업 중인 폴더 경로
            pdf_files: 전체 PDF 파일 리스트
            completed_files: 완료된 파일 리스트
            current_index: 현재 작업 중인 파일 인덱스
            
        Returns:
            tuple[bool, str]: (성공 여부, 메시지)
        """
        try:
            data = {
                'folder_path': folder_path,
                'last_updated': datetime.now().isoformat(),
                'total_files': len(pdf_files),
                'completed_count': len(completed_files),
                'current_index': current_index,
                'pdf_files': [os.path.basename(f) for f in pdf_files],
                'completed_files': completed_files,
            }
            
            with open(self.progress_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            
            return True, "진행상황 저장 완료"
            
        except Exception as e:
            return False, f"진행상황 저장 실패: {str(e)}"
    
    def load_progress(self) -> tuple[bool, dict, str]:
        """
        작업 진행상황 로드
        
        Returns:
            tuple[bool, dict, str]: (성공 여부, 진행상황 데이터, 메시지)
        """
        try:
            if not os.path.exists(self.progress_file):
                return True, {}, "진행상황 파일 없음"
            
            with open(self.progress_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            return True, data, "진행상황 로드 완료"
            
        except Exception as e:
            return False, {}, f"진행상황 로드 실패: {str(e)}"
    
    def clear_progress(self) -> tuple[bool, str]:
        """
        작업 진행상황 삭제
        
        Returns:
            tuple[bool, str]: (성공 여부, 메시지)
        """
        try:
            if os.path.exists(self.progress_file):
                os.remove(self.progress_file)
                return True, "진행상황 삭제 완료"
            else:
                return True, "진행상황 파일 없음"
                
        except Exception as e:
            return False, f"진행상황 삭제 실패: {str(e)}"


class MaskDataManager:
    """마스킹 데이터 관리 클래스"""
    
    def __init__(self) -> None:
        # 프로젝트 루트 경로
        self.project_root = os.path.dirname(os.path.abspath(__file__))
        self.masks_dir = os.path.join(self.project_root, "masks_data")
        
        # 마스킹 데이터 폴더 생성
        os.makedirs(self.masks_dir, exist_ok=True)
    
    def get_mask_file_path(self) -> str:
        """
        일자별 마스킹 데이터 파일 경로 반환
        
        Returns:
            str: 마스킹 데이터 파일 경로 (masks_data/mask_data_YYYYMMDD.json)
        """
        today_str = datetime.now().strftime("%Y%m%d")
        mask_filename = f"mask_data_{today_str}.json"
        return os.path.join(self.masks_dir, mask_filename)
    
    def save_masks(self, pdf_path: str, masks: list[MaskEntry]) -> tuple[bool, str]:
        """
        마스킹 데이터를 일자별 JSON 파일에 추가 저장
        
        Args:
            pdf_path: PDF 파일 경로
            masks: 마스킹 데이터 리스트
            
        Returns:
            tuple[bool, str]: (성공 여부, 메시지)
        """
        try:
            mask_file = self.get_mask_file_path()
            pdf_filename = os.path.basename(pdf_path)
            
            # 기존 데이터 로드
            if os.path.exists(mask_file):
                with open(mask_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
            else:
                data = {
                    'date': datetime.now().strftime("%Y-%m-%d"),
                    'files': []
                }
            
            # MaskEntry를 딕셔너리로 변환
            masks_data = []
            for mask in masks:
                masks_data.append({
                    'page_index': mask.page_index,
                    'rect': {
                        'x0': mask.rect.x0,
                        'y0': mask.rect.y0,
                        'x1': mask.rect.x1,
                        'y1': mask.rect.y1,
                    },
                    'note': mask.note,
                })
            
            # 파일별 데이터 구성
            file_data = {
                'pdf_file': pdf_filename,
                'saved_at': datetime.now().isoformat(),
                'mask_count': len(masks),
                'masks': masks_data,
            }
            
            # 기존 파일 데이터가 있으면 업데이트, 없으면 추가
            file_found = False
            for i, file_entry in enumerate(data['files']):
                if file_entry['pdf_file'] == pdf_filename:
                    data['files'][i] = file_data
                    file_found = True
                    break
            
            if not file_found:
                data['files'].append(file_data)
            
            # JSON으로 저장
            with open(mask_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            
            return True, mask_file
            
        except Exception as e:
            return False, f"마스킹 데이터 저장 실패: {str(e)}"
    
    def load_masks(self, pdf_path: str) -> tuple[bool, list[MaskEntry], str]:
        """
        일자별 JSON 파일에서 특정 PDF의 마스킹 데이터 로드
        
        Args:
            pdf_path: PDF 파일 경로
            
        Returns:
            tuple[bool, list[MaskEntry], str]: (성공 여부, 마스킹 리스트, 메시지)
        """
        try:
            mask_file = self.get_mask_file_path()
            pdf_filename = os.path.basename(pdf_path)
            
            if not os.path.exists(mask_file):
                return True, [], "오늘 날짜의 마스킹 데이터 없음"
            
            with open(mask_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # 해당 PDF 파일의 데이터 찾기
            file_data = None
            for file_entry in data.get('files', []):
                if file_entry['pdf_file'] == pdf_filename:
                    file_data = file_entry
                    break
            
            if not file_data:
                return True, [], f"{pdf_filename}의 마스킹 데이터 없음"
            
            # 딕셔너리를 MaskEntry로 변환
            masks = []
            for mask_data in file_data.get('masks', []):
                rect_data = mask_data['rect']
                rect = fitz.Rect(
                    rect_data['x0'],
                    rect_data['y0'],
                    rect_data['x1'],
                    rect_data['y1']
                )
                mask = MaskEntry(
                    page_index=mask_data['page_index'],
                    rect=rect,
                    note=mask_data.get('note', '')
                )
                masks.append(mask)
            
            return True, masks, f"{len(masks)}개의 마스킹 데이터 로드 완료"
            
        except Exception as e:
            return False, [], f"마스킹 데이터 로드 실패: {str(e)}"
    
    def delete_masks(self, pdf_path: str) -> tuple[bool, str]:
        """
        일자별 JSON 파일에서 특정 PDF의 마스킹 데이터 삭제
        
        Args:
            pdf_path: PDF 파일 경로
            
        Returns:
            tuple[bool, str]: (성공 여부, 메시지)
        """
        try:
            mask_file = self.get_mask_file_path()
            pdf_filename = os.path.basename(pdf_path)
            
            if not os.path.exists(mask_file):
                return True, "마스킹 데이터 파일 없음"
            
            with open(mask_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # 해당 PDF 파일 데이터 제거
            original_count = len(data.get('files', []))
            data['files'] = [f for f in data.get('files', []) if f['pdf_file'] != pdf_filename]
            
            if len(data['files']) < original_count:
                # 데이터 저장
                with open(mask_file, 'w', encoding='utf-8') as f:
                    json.dump(data, f, indent=2, ensure_ascii=False)
                return True, "마스킹 데이터 삭제 완료"
            else:
                return True, "해당 파일의 마스킹 데이터 없음"
                
        except Exception as e:
            return False, f"마스킹 데이터 삭제 실패: {str(e)}"


class PdfDocumentManager:
    """PDF 문서 관리 클래스"""

    def __init__(self) -> None:
        self.doc: Optional[fitz.Document] = None
        self.file_path: Optional[str] = None

    def load_pdf(self, path: str) -> None:
        """PDF 파일 로드"""
        try:
            # 기존 문서가 있다면 닫기
            if self.doc is not None:
                self.doc.close()
                self.doc = None

            # 새 문서 열기
            self.doc = fitz.open(path)
            self.file_path = path
            
        except Exception as e:
            self.doc = None
            self.file_path = None
            raise Exception(f"PDF 파일을 열 수 없습니다: {str(e)}")

    def get_page_count(self) -> int:
        """전체 페이지 수 반환"""
        if self.doc is None:
            return 0
        return len(self.doc)

    def get_page_pixmap(self, page_index: int, zoom: float = 1.5) -> Optional[QPixmap]:
        """지정된 페이지를 QPixmap으로 렌더링"""
        if self.doc is None:
            return None

        try:
            # 페이지 범위 체크
            if page_index < 0 or page_index >= len(self.doc):
                return None

            # 페이지 로드
            page = self.doc.load_page(page_index)

            # 확대/축소 매트릭스 적용하여 렌더링
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat)

            # PyMuPDF Pixmap을 QImage로 변환
            img_data = pix.samples
            img = QImage(
                img_data,
                pix.width,
                pix.height,
                pix.stride,
                QImage.Format.Format_RGB888
            )

            # QImage를 QPixmap으로 변환
            return QPixmap.fromImage(img)

        except Exception as e:
            print(f"페이지 렌더링 오류: {str(e)}")
            return None

    def apply_masks_and_save(self, masks: list[MaskEntry]) -> None:
        """
        마스킹을 PDF에 적용하고 저장
        
        Args:
            masks: 적용할 마스킹 정보 리스트
            
        Raises:
            Exception: 문서가 없거나 저장 중 오류 발생 시
        """
        if self.doc is None or self.file_path is None:
            raise Exception("열린 PDF 문서가 없습니다.")
        
        try:
            # 페이지별로 마스크 적용
            for page_num in range(len(self.doc)):
                # 현재 페이지의 마스크만 필터링
                page_masks = [m for m in masks if m.page_index == page_num]
                
                if not page_masks:
                    continue
                
                # 페이지 로드
                page = self.doc.load_page(page_num)
                
                # 각 마스크에 대해 redaction annotation 추가
                for mask in page_masks:
                    # 흰색으로 마스킹 (1, 1, 1) = RGB white
                    page.add_redact_annot(mask.rect, fill=(1, 1, 1))
                
                # 페이지별로 redaction 적용
                page.apply_redactions()
                
                print(f"페이지 {page_num + 1}에 {len(page_masks)}개의 마스크 적용 완료")
            
            print("모든 페이지의 Redaction 적용 완료")
            
            # 파일 저장 (incremental 저장)
            self.doc.save(
                self.file_path,
                incremental=True,
                encryption=fitz.PDF_ENCRYPT_KEEP
            )
            print(f"파일 저장 완료: {self.file_path}")
            
        except Exception as e:
            error_msg = str(e)
            # Permission denied 오류 확인
            if "Permission denied" in error_msg or "permission denied" in error_msg.lower():
                raise Exception(
                    "파일 저장 권한이 없습니다.\n\n"
                    "PDF 파일이 다른 프로그램(Adobe Reader, 웹 브라우저 등)에서 열려있을 수 있습니다.\n"
                    "해당 프로그램을 종료한 후 다시 시도해주세요.\n\n"
                    f"파일 경로: {self.file_path}"
                )
            else:
                raise Exception(f"마스킹 저장 중 오류 발생: {error_msg}")

    def close(self) -> None:
        """문서 닫기"""
        if self.doc is not None:
            self.doc.close()
            self.doc = None
        self.file_path = None


class MainWindow(QMainWindow):
    """메인 윈도우 클래스"""

    def __init__(self) -> None:
        super().__init__()
        
        # PDF 문서 관리자
        self.pdf_manager = PdfDocumentManager()
        self.current_page_index: int = 0
        
        # 마스킹 데이터 관리자
        self.mask_data_manager = MaskDataManager()
        
        # 진행상황 관리자
        self.progress_manager = ProgressManager()
        
        # 로그 관리자
        self.log_manager = LogManager()
        self.log_manager.log_app_start()
        
        # 마스킹 데이터 저장
        self.masks: list[MaskEntry] = []
        
        # 폴더 내 PDF 파일 목록
        self.pdf_files: list[str] = []
        self.current_pdf_index: int = -1
        self.completed_files: list[str] = []  # 완료된 파일 리스트
        self.current_folder_path: str = ""  # 현재 작업 중인 폴더 경로
        
        # 백업 설정
        self.backup_enabled: bool = True  # 백업 활성화 여부 (기본값: 활성화)
        
        self.init_ui()
        self.setup_menu()
        self.setup_toolbar()
        self.setup_shortcuts()
        self.setup_statusbar()

    def init_ui(self) -> None:
        """UI 초기화"""
        self.setWindowTitle("PDF Mask - PDF 마스킹 프로그램")
        self.setGeometry(100, 100, 1400, 900)
        
        # Dock Widget 도킹 옵션 설정 (최대화 상태에서도 도킹 가능하도록)
        self.setDockOptions(
            QMainWindow.DockOption.AllowNestedDocks |      # 중첩 도킹 허용
            QMainWindow.DockOption.AllowTabbedDocks |      # 탭 도킹 허용
            QMainWindow.DockOption.AnimatedDocks           # 애니메이션 효과
        )

        # 중앙 위젯 설정
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # 메인 레이아웃 (수평)
        main_layout = QHBoxLayout()
        central_widget.setLayout(main_layout)

        # 가운데: PDF 페이지 표시 (스크롤 가능한 뷰)
        self.scrollable_pdf_view = ScrollablePdfView(self)
        self.pdf_view = self.scrollable_pdf_view.pdf_view
        
        # 시그널 연결
        self.pdf_view.maskCreated.connect(self.on_mask_created)
        
        main_layout.addWidget(self.scrollable_pdf_view)
        
        # 왼쪽: 마스킹 리스트 (Dock Widget)
        self.mask_list = QTableWidget()
        self.setup_mask_table()
        self.mask_dock_widget = QDockWidget("마스킹 리스트", self)
        self.mask_dock_widget.setWidget(self.mask_list)
        self.mask_dock_widget.setMinimumWidth(250)
        
        # Dock 기능 설정
        self.mask_dock_widget.setFeatures(
            QDockWidget.DockWidgetFeature.DockWidgetMovable |     # 이동 가능
            QDockWidget.DockWidgetFeature.DockWidgetFloatable |   # 분리 가능
            QDockWidget.DockWidgetFeature.DockWidgetClosable      # 닫기 가능
        )
        
        # 도킹 영역 설정 (모든 영역 허용)
        self.mask_dock_widget.setAllowedAreas(
            Qt.DockWidgetArea.LeftDockWidgetArea |
            Qt.DockWidgetArea.RightDockWidgetArea |
            Qt.DockWidgetArea.TopDockWidgetArea |
            Qt.DockWidgetArea.BottomDockWidgetArea
        )
        
        self.addDockWidget(Qt.DockWidgetArea.LeftDockWidgetArea, self.mask_dock_widget)

        # 오른쪽: PDF 파일 목록 (Dock Widget)
        self.pdf_file_list = QListWidget()
        self.pdf_file_list.itemDoubleClicked.connect(self.on_pdf_list_double_clicked)
        self.pdf_dock_widget = QDockWidget("PDF 파일 목록", self)
        self.pdf_dock_widget.setWidget(self.pdf_file_list)
        self.pdf_dock_widget.setMinimumWidth(200)
        
        # Dock 기능 설정
        self.pdf_dock_widget.setFeatures(
            QDockWidget.DockWidgetFeature.DockWidgetMovable |     # 이동 가능
            QDockWidget.DockWidgetFeature.DockWidgetFloatable |   # 분리 가능
            QDockWidget.DockWidgetFeature.DockWidgetClosable      # 닫기 가능
        )
        
        # 도킹 영역 설정 (모든 영역 허용)
        self.pdf_dock_widget.setAllowedAreas(
            Qt.DockWidgetArea.LeftDockWidgetArea |
            Qt.DockWidgetArea.RightDockWidgetArea |
            Qt.DockWidgetArea.TopDockWidgetArea |
            Qt.DockWidgetArea.BottomDockWidgetArea
        )
        
        self.addDockWidget(Qt.DockWidgetArea.RightDockWidgetArea, self.pdf_dock_widget)

    def setup_mask_table(self) -> None:
        """마스킹 리스트 테이블 설정"""
        # 컬럼 설정
        self.mask_list.setColumnCount(2)
        self.mask_list.setHorizontalHeaderLabels(["페이지", "메모"])
        
        # 크기 설정
        self.mask_list.setMinimumWidth(250)
        self.mask_list.setMaximumWidth(400)
        
        # 행 헤더 숨김
        self.mask_list.verticalHeader().setVisible(False)
        
        # 컬럼 크기 조절
        self.mask_list.setColumnWidth(0, 60)  # 페이지 컬럼
        self.mask_list.horizontalHeader().setStretchLastSection(True)  # 메모 컬럼 자동 확장
        
        # 선택 모드: 행 단위 선택
        self.mask_list.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        
        # 편집 트리거: 더블클릭
        self.mask_list.setEditTriggers(
            QTableWidget.EditTrigger.DoubleClicked | 
            QTableWidget.EditTrigger.EditKeyPressed
        )
        
        # 시그널 연결
        self.mask_list.itemChanged.connect(self.on_mask_item_changed)

    def setup_menu(self) -> None:
        """메뉴바 설정"""
        menubar = self.menuBar()

        # 파일 메뉴
        file_menu = menubar.addMenu("파일(&F)")

        # PDF 열기 액션
        open_pdf_action = QAction("PDF 열기...", self)
        open_pdf_action.setShortcut(QKeySequence("Ctrl+O"))
        open_pdf_action.triggered.connect(self.open_pdf)
        file_menu.addAction(open_pdf_action)

        # 폴더 열기 액션
        open_folder_action = QAction("폴더 열기...", self)
        open_folder_action.setShortcut(QKeySequence("Ctrl+Shift+O"))
        open_folder_action.triggered.connect(self.open_folder)
        file_menu.addAction(open_folder_action)

        # 마스킹 내역 엑셀 저장 액션
        export_excel_action = QAction("마스킹 내역 엑셀 저장...", self)
        export_excel_action.triggered.connect(self.export_masks_to_excel)
        file_menu.addAction(export_excel_action)

        file_menu.addSeparator()

        # 종료 액션
        exit_action = QAction("종료", self)
        exit_action.setShortcut(QKeySequence("Ctrl+Q"))
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)
        
        # 설정 메뉴
        settings_menu = menubar.addMenu("설정(&S)")
        
        # 백업 활성화/비활성화 액션 (체크 가능)
        self.backup_toggle_action = QAction("PDF 백업 활성화", self)
        self.backup_toggle_action.setCheckable(True)
        self.backup_toggle_action.setChecked(self.backup_enabled)
        self.backup_toggle_action.triggered.connect(self.toggle_backup)
        settings_menu.addAction(self.backup_toggle_action)
        
        # 보기 메뉴
        view_menu = menubar.addMenu("보기(&V)")
        
        # 마스킹 리스트 토글 액션
        self.toggle_mask_list_action = self.mask_dock_widget.toggleViewAction()
        self.toggle_mask_list_action.setText("마스킹 리스트")
        view_menu.addAction(self.toggle_mask_list_action)
        
        # PDF 파일 목록 토글 액션
        self.toggle_pdf_list_action = self.pdf_dock_widget.toggleViewAction()
        self.toggle_pdf_list_action.setText("PDF 파일 목록")
        view_menu.addAction(self.toggle_pdf_list_action)

        # 도움말 메뉴
        help_menu = menubar.addMenu("도움말(&H)")

        # 단축키 안내
        shortcuts_action = QAction("단축키 안내", self)
        shortcuts_action.triggered.connect(self.show_shortcuts_help)
        help_menu.addAction(shortcuts_action)

        # 사용 방법
        usage_action = QAction("사용 방법", self)
        usage_action.triggered.connect(self.show_usage_help)
        help_menu.addAction(usage_action)

        # 정보
        about_action = QAction("정보", self)
        about_action.triggered.connect(self.show_about_dialog)
        help_menu.addAction(about_action)

    def setup_toolbar(self) -> None:
        """툴바 설정"""
        toolbar = QToolBar("메인 툴바")
        toolbar.setIconSize(QSize(24, 24))
        self.addToolBar(toolbar)

        # PDF 열기 버튼
        file_icon = self.style().standardIcon(QStyle.StandardPixmap.SP_FileIcon)
        open_pdf_action = QAction(file_icon, "PDF 열기", self)
        open_pdf_action.triggered.connect(self.open_pdf)
        toolbar.addAction(open_pdf_action)

        # 폴더 열기 버튼
        folder_icon = self.style().standardIcon(QStyle.StandardPixmap.SP_DirOpenIcon)
        open_folder_action = QAction(folder_icon, "폴더 열기", self)
        open_folder_action.triggered.connect(self.open_folder)
        toolbar.addAction(open_folder_action)

    def setup_shortcuts(self) -> None:
        """키보드 단축키 설정"""
        from PyQt6.QtGui import QShortcut

        # 다음 페이지 (오른쪽 방향키)
        self.shortcut_next_page = QShortcut(QKeySequence(Qt.Key.Key_Right), self)
        self.shortcut_next_page.activated.connect(self.go_next_page)

        # 이전 페이지 (왼쪽 방향키)
        self.shortcut_prev_page = QShortcut(QKeySequence(Qt.Key.Key_Left), self)
        self.shortcut_prev_page.activated.connect(self.go_prev_page)

        # 다음 페이지 (PageDown)
        self.shortcut_page_down = QShortcut(QKeySequence(Qt.Key.Key_PageDown), self)
        self.shortcut_page_down.activated.connect(self.go_next_page)

        # 이전 페이지 (PageUp)
        self.shortcut_page_up = QShortcut(QKeySequence(Qt.Key.Key_PageUp), self)
        self.shortcut_page_up.activated.connect(self.go_prev_page)

        # 저장 (Ctrl+S)
        self.shortcut_save = QShortcut(QKeySequence("Ctrl+S"), self)
        self.shortcut_save.activated.connect(self.save_masks)
        
        # 삭제 (Del)
        self.shortcut_delete = QShortcut(QKeySequence(Qt.Key.Key_Delete), self)
        self.shortcut_delete.activated.connect(self.delete_selected_mask)

        # 확대 (+)
        self.shortcut_zoom_in = QShortcut(QKeySequence(Qt.Key.Key_Plus), self)
        self.shortcut_zoom_in.activated.connect(self.zoom_in)

        # 축소 (-)
        self.shortcut_zoom_out = QShortcut(QKeySequence(Qt.Key.Key_Minus), self)
        self.shortcut_zoom_out.activated.connect(self.zoom_out)

    def setup_statusbar(self) -> None:
        """상태바 설정"""
        self.statusBar().showMessage("준비 (Ctrl + 드래그로 마스킹 영역 선택)")

    def update_page_view(self) -> None:
        """현재 페이지를 화면에 표시"""
        # 문서가 없으면 리턴
        if self.pdf_manager.doc is None:
            return

        try:
            # 현재 페이지 객체 가져오기
            page = self.pdf_manager.doc[self.current_page_index]
            
            # 페이지 실제 크기
            page_width = page.rect.width
            page_height = page.rect.height
            
            # 줌 레벨 적용
            zoom = self.scrollable_pdf_view.zoom_level * 1.5
            
            # 페이지 렌더링
            pixmap = self.pdf_manager.get_page_pixmap(self.current_page_index, zoom)
            
            if pixmap is not None:
                # PdfPageView에 페이지 설정 (마스킹 영역 정보도 전달)
                self.pdf_view.set_page(
                    self.current_page_index,
                    pixmap,
                    page_width,
                    page_height,
                    self.masks
                )
                
                # 상태바 업데이트
                total_pages = self.pdf_manager.get_page_count()
                zoom_percent = int(self.scrollable_pdf_view.zoom_level * 100)
                self.statusBar().showMessage(
                    f"페이지: {self.current_page_index + 1} / {total_pages} | "
                    f"확대: {zoom_percent}% | "
                    f"(Ctrl + 드래그: 마스킹, Ctrl + 휠: 확대/축소)"
                )
            else:
                self.statusBar().showMessage("오류: 페이지 렌더링 실패")
                
        except Exception as e:
            print(f"페이지 표시 오류: {str(e)}")
            self.statusBar().showMessage(f"오류: {str(e)}")

    def zoom_in(self) -> None:
        """PDF 확대"""
        self.scrollable_pdf_view.zoom_in()

    def zoom_out(self) -> None:
        """PDF 축소"""
        self.scrollable_pdf_view.zoom_out()

    def show_shortcuts_help(self) -> None:
        """단축키 안내 다이얼로그 표시"""
        shortcuts_text = (
            "<b>[파일 관련]</b><br>"
            "<b>Ctrl+O</b> : PDF 열기<br>"
            "<b>Ctrl+Shift+O</b> : 폴더 열기<br>"
            "<b>Ctrl+S</b> : 마스킹 저장<br><br>"
            "<b>[페이지 이동]</b><br>"
            "<b>→ / PageDown</b> : 다음 페이지<br>"
            "<b>← / PageUp</b> : 이전 페이지<br><br>"
            "<b>[마스킹]</b><br>"
            "<b>Ctrl+드래그</b> : 마스킹 영역 선택<br>"
            "<b>Del</b> : 선택된 마스킹 삭제<br><br>"
            "<b>[기타]</b><br>"
            "<b>Ctrl+Q</b> : 프로그램 종료"
        )
        QMessageBox.information(self, "단축키 안내", shortcuts_text)

    def show_usage_help(self) -> None:
        """사용 방법 안내 다이얼로그 표시"""
        usage_text = (
            "1. 상단 툴바 또는 파일 메뉴에서 PDF 또는 폴더를 엽니다.\n"
            "2. 좌측 '마스킹 리스트' 패널은 생성된 마스킹 정보를 보여줍니다.\n"
            "3. 중앙 PDF 화면에서 Ctrl 키를 누른 상태로 마우스를 드래그하여 마스킹 영역을 선택합니다.\n"
            "4. 우측 'PDF 파일 목록'에서 다른 PDF를 더블클릭하여 전환할 수 있습니다.\n"
            "5. 마스킹이 완료되면 Ctrl+S로 마스킹을 적용하고 저장합니다.\n"
        )
        QMessageBox.information(self, "사용 방법", usage_text)

    def show_about_dialog(self) -> None:
        """프로그램 정보 다이얼로그 표시"""
        about_text = (
            "PDF Mask\n\n"
            "PDF 문서의 민감한 정보를 마스킹하기 위한 PyQt6 기반 도구입니다.\n"
            "PyMuPDF를 사용하여 PDF를 렌더링하고, Redaction 기능으로 내용을 영구적으로 가립니다.\n\n"
            "버전: 1.0\n"
            "제작: (작성자 정보 기입 예정)"
        )
        QMessageBox.information(self, "정보", about_text)

    def reload_current_pdf(self) -> None:
        """현재 PDF 파일을 다시 로드"""
        if self.pdf_manager.file_path is None:
            return
        
        try:
            current_path = self.pdf_manager.file_path
            current_page = self.current_page_index
            
            # PDF 재로드
            self.pdf_manager.load_pdf(current_path)
            
            # 현재 페이지로 이동 (페이지 범위 체크)
            page_count = self.pdf_manager.get_page_count()
            if current_page >= page_count:
                self.current_page_index = max(0, page_count - 1)
            else:
                self.current_page_index = current_page
            
            # 페이지 표시
            self.update_page_view()
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "오류",
                f"PDF 파일을 다시 로드할 수 없습니다.\n\n{str(e)}"
            )

    def clear_masks(self) -> None:
        """마스킹 데이터 및 테이블 초기화"""
        self.masks.clear()
        self.mask_list.setRowCount(0)

    def load_pdf_from_path(self, file_path: str) -> None:
        """
        지정된 경로의 PDF 파일 로드
        
        Args:
            file_path: 로드할 PDF 파일 경로
        """
        try:
            # PDF 로드
            self.pdf_manager.load_pdf(file_path)
            
            # 마스킹 데이터 초기화
            self.clear_masks()
            
            # 저장된 마스킹 데이터 로드 시도
            success, loaded_masks, msg = self.mask_data_manager.load_masks(file_path)
            if success and loaded_masks:
                self.masks = loaded_masks
                # 테이블에 마스킹 데이터 표시
                for mask in self.masks:
                    row = self.mask_list.rowCount()
                    self.mask_list.insertRow(row)
                    
                    # 페이지 컬럼
                    page_item = QTableWidgetItem(str(mask.page_index + 1))
                    page_item.setFlags(page_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    self.mask_list.setItem(row, 0, page_item)
                    
                    # 메모 컬럼
                    note_item = QTableWidgetItem(mask.note)
                    self.mask_list.setItem(row, 1, note_item)
                    
                    page_item.setData(Qt.ItemDataRole.UserRole, row)
                
                print(f"마스킹 데이터 로드: {msg}")
            
            # 첫 페이지로 이동
            self.current_page_index = 0
            
            # 줌 레벨 초기화
            self.scrollable_pdf_view.zoom_level = 1.0
            
            # 페이지 표시
            self.update_page_view()
            
            # 창 제목에 파일명 표시
            filename = os.path.basename(file_path)
            self.setWindowTitle(f"PDF Mask - {filename}")
            
            # 로그 기록
            self.log_manager.log_pdf_open(file_path)
            print(f"PDF 로드 성공: {file_path}")
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "오류",
                f"PDF 파일을 열 수 없습니다.\n\n{str(e)}"
            )
            print(f"PDF 로드 실패: {str(e)}")

    def load_pdf_from_list(self, index: int) -> None:
        """
        PDF 파일 목록에서 지정된 인덱스의 PDF 로드
        
        Args:
            index: pdf_files 리스트의 인덱스
        """
        if index < 0 or index >= len(self.pdf_files):
            return
        
        self.current_pdf_index = index
        file_path = self.pdf_files[index]
        self.load_pdf_from_path(file_path)
        
        # 리스트에서 현재 파일 하이라이트
        self.pdf_file_list.setCurrentRow(index)

    def delete_selected_mask(self) -> None:
        """선택된 마스킹 항목 삭제 (Del 키)"""
        selected_rows = self.mask_list.selectionModel().selectedRows()
        
        if not selected_rows:
            return
        
        # 역순으로 삭제 (인덱스 변경 방지)
        rows_to_delete = sorted([index.row() for index in selected_rows], reverse=True)
        
        for row in rows_to_delete:
            if 0 <= row < len(self.masks):
                # 데이터에서 삭제
                del self.masks[row]
                # 테이블에서 삭제
                self.mask_list.removeRow(row)
        
        # 화면 갱신 (마스킹 영역 업데이트)
        self.update_page_view()
        
        print(f"{len(rows_to_delete)}개의 마스킹 항목 삭제됨")

    def toggle_backup(self) -> None:
        """백업 활성화/비활성화 토글"""
        self.backup_enabled = self.backup_toggle_action.isChecked()
        status = "활성화" if self.backup_enabled else "비활성화"
        print(f"PDF 백업 {status}")

    def backup_current_pdf(self) -> tuple[bool, str]:
        """
        현재 PDF 파일을 백업 폴더에 복사
        
        Returns:
            tuple[bool, str]: (성공 여부, 메시지 또는 백업 경로)
        """
        if self.pdf_manager.file_path is None:
            return False, "백업할 PDF 파일이 없습니다."
        
        try:
            # 백업 폴더 구조: 프로젝트최상단/backup/YYYYMMDD/원본파일명.pdf
            pdf_path = self.pdf_manager.file_path
            pdf_filename = os.path.basename(pdf_path)
            
            # 프로젝트 최상단 경로 (main.py가 있는 위치)
            project_root = os.path.dirname(os.path.abspath(__file__))
            
            # 백업 폴더 생성
            today_str = datetime.now().strftime("%Y%m%d")
            backup_base = os.path.join(project_root, "backup")
            backup_dir = os.path.join(backup_base, today_str)
            
            # 폴더가 없으면 생성
            os.makedirs(backup_dir, exist_ok=True)
            
            # 백업 파일 경로
            backup_path = os.path.join(backup_dir, pdf_filename)
            
            # 이미 백업된 파일이 있는지 확인
            if os.path.exists(backup_path):
                # 같은 파일이면 백업 스킵
                return True, f"이미 백업된 파일: {backup_path}"
            
            # 파일 복사
            shutil.copy2(pdf_path, backup_path)
            
            return True, backup_path
            
        except Exception as e:
            return False, f"백업 실패: {str(e)}"

    def export_masks_to_excel(self) -> None:
        """마스킹 작업 내역을 엑셀 파일로 저장"""
        # 저장할 마스크가 없는 경우
        if not self.masks:
            QMessageBox.information(self, "알림", "저장할 마스킹 내역이 없습니다.")
            return

        # 현재 PDF 파일 정보가 없는 경우
        if self.pdf_manager.file_path is None:
            QMessageBox.warning(self, "경고", "열려 있는 PDF 파일 정보가 없습니다.")
            return

        # openpyxl 지연 로딩
        try:
            from openpyxl import Workbook, load_workbook
        except ImportError:
            QMessageBox.critical(
                self,
                "오류",
                "openpyxl 패키지가 설치되어 있지 않습니다.\n\n"
                "다음 명령으로 설치 후 다시 시도해주세요.\n"
                "pip install openpyxl"
            )
            return

        from datetime import datetime

        try:
            # 저장 경로 및 파일명 구성 (현재 PDF와 같은 폴더)
            pdf_path = self.pdf_manager.file_path
            base_dir = os.path.dirname(pdf_path) or os.getcwd()
            today_str = datetime.now().strftime("%Y%m%d")
            filename = f"마스킹_작업내역_{today_str}.xlsx"
            save_path = os.path.join(base_dir, filename)

            # 워크북 생성 또는 로드
            if os.path.exists(save_path):
                wb = load_workbook(save_path)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "마스킹 내역"
                # 헤더 행
                ws.append(["작업일시", "PDF 파일명", "페이지", "마스킹 영역 좌표", "메모"])

            pdf_name = os.path.basename(pdf_path)

            # 마스크 데이터 추가
            for mask in self.masks:
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                page_number = mask.page_index + 1
                rect = mask.rect
                coords = f"({rect.x0:.2f}, {rect.y0:.2f}, {rect.x1:.2f}, {rect.y1:.2f})"
                note = mask.note or ""

                ws.append([timestamp, pdf_name, page_number, coords, note])

            # 파일 저장 (별도 완료 메시지는 표시하지 않음)
            wb.save(save_path)

        except Exception as e:
            QMessageBox.critical(
                self,
                "저장 오류",
                f"엑셀 파일 저장 중 오류가 발생했습니다.\n\n{str(e)}"
            )

    # 메뉴/툴바 액션 슬롯
    def open_pdf(self) -> None:
        """PDF 파일 열기"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "PDF 파일 열기", "", "PDF Files (*.pdf);;All Files (*)"
        )
        
        if file_path:
            self.load_pdf_from_path(file_path)
            
            # 단일 파일 열기이므로 pdf_files 리스트는 이 파일만 포함
            self.pdf_files = [file_path]
            self.current_pdf_index = 0
            
            # PDF 파일 리스트 업데이트
            self.pdf_file_list.clear()
            self.pdf_file_list.addItem(os.path.basename(file_path))
            self.pdf_file_list.setCurrentRow(0)

    def open_folder(self) -> None:
        """폴더 열기 및 PDF 파일 목록 로드"""
        folder_path = QFileDialog.getExistingDirectory(self, "폴더 선택")
        
        if not folder_path:
            return
        
        print(f"선택된 폴더: {folder_path}")
        
        # 폴더 내 PDF 파일 검색
        pdf_files = []
        try:
            for filename in os.listdir(folder_path):
                if filename.lower().endswith('.pdf'):
                    full_path = os.path.join(folder_path, filename)
                    pdf_files.append(full_path)
        except Exception as e:
            QMessageBox.critical(
                self,
                "오류",
                f"폴더를 읽을 수 없습니다.\n\n{str(e)}"
            )
            return
        
        # PDF 파일이 없는 경우
        if not pdf_files:
            QMessageBox.information(
                self,
                "알림",
                "선택한 폴더에 PDF 파일이 없습니다."
            )
            return
        
        # 파일명으로 정렬
        pdf_files.sort()
        
        # PDF 파일 목록 저장
        self.pdf_files = pdf_files
        self.current_folder_path = folder_path
        self.completed_files = []  # 완료 목록 초기화
        
        # 이전 진행상황 확인
        success, progress_data, msg = self.progress_manager.load_progress()
        if success and progress_data and progress_data.get('folder_path') == folder_path:
            # 이전에 작업하던 폴더와 동일한 경우
            reply = QMessageBox.question(
                self,
                "진행상황 복구",
                f"이전 작업 진행상황이 있습니다.\n\n"
                f"완료: {progress_data.get('completed_count', 0)}/{progress_data.get('total_files', 0)}\n"
                f"마지막 작업: {progress_data.get('last_updated', '')}\n\n"
                f"이어서 작업하시겠습니까?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.Yes
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                # 진행상황 복구
                self.completed_files = progress_data.get('completed_files', [])
                start_index = progress_data.get('current_index', 0)
                
                # 이미 완료된 파일은 비활성화 표시 (선택적)
                print(f"진행상황 복구: {len(self.completed_files)}개 파일 완료")
        
        # 리스트 위젯 업데이트
        self.pdf_file_list.clear()
        for file_path in pdf_files:
            filename = os.path.basename(file_path)
            item_text = filename
            # 완료된 파일 표시
            if filename in self.completed_files:
                item_text = f"✓ {filename}"
            self.pdf_file_list.addItem(item_text)
        
        # 로그 기록
        self.log_manager.log_folder_open(folder_path, len(pdf_files))
        print(f"총 {len(pdf_files)}개의 PDF 파일 발견")
        
        # 첫 번째 PDF 또는 복구된 위치에서 시작
        if pdf_files:
            if success and progress_data and progress_data.get('folder_path') == folder_path:
                start_index = progress_data.get('current_index', 0)
                # 인덱스 범위 체크
                if 0 <= start_index < len(pdf_files):
                    self.load_pdf_from_list(start_index)
                else:
                    self.load_pdf_from_list(0)
            else:
                self.load_pdf_from_list(0)

    # PDF 리스트 이벤트
    def on_pdf_list_double_clicked(self, item) -> None:
        """PDF 파일 리스트에서 더블클릭 이벤트"""
        row = self.pdf_file_list.row(item)
        self.load_pdf_from_list(row)

    # 키보드 단축키 슬롯
    def go_next_page(self) -> None:
        """다음 페이지로 이동"""
        if self.pdf_manager.doc is None:
            return
        
        page_count = self.pdf_manager.get_page_count()
        if self.current_page_index < page_count - 1:
            self.current_page_index += 1
            self.update_page_view()
            print(f"다음 페이지: {self.current_page_index + 1}")

    def go_prev_page(self) -> None:
        """이전 페이지로 이동"""
        if self.pdf_manager.doc is None:
            return
        
        if self.current_page_index > 0:
            self.current_page_index -= 1
            self.update_page_view()
            print(f"이전 페이지: {self.current_page_index + 1}")

    def save_masks(self) -> None:
        """
        마스킹 정보 저장 (Ctrl+S)
        
        마스크가 있으면 PDF에 적용하여 저장
        마스크가 없으면 다음 파일로 이동 확인
        """
        # PDF가 열려있지 않으면 무시
        if self.pdf_manager.doc is None:
            return
        
        # (1) 마스크가 1개 이상 있는 경우
        if len(self.masks) > 0:
            reply = QMessageBox.question(
                self,
                "마스킹 저장",
                "현재 PDF 파일의 마스킹 작업 내용을 저장하시겠습니까?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.Yes
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                try:
                    # 백업 수행 (활성화된 경우)
                    if self.backup_enabled:
                        backup_success, backup_msg = self.backup_current_pdf()
                        
                        if not backup_success:
                            # 백업 실패 시 사용자에게 확인
                            retry_reply = QMessageBox.warning(
                                self,
                                "백업 실패",
                                f"{backup_msg}\n\n백업 없이 계속 진행하시겠습니까?",
                                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                                QMessageBox.StandardButton.No
                            )
                            
                            if retry_reply == QMessageBox.StandardButton.No:
                                return  # 저장 취소
                        else:
                            print(f"백업 완료: {backup_msg}")
                    
                    # 마스킹 데이터 JSON 저장 (PDF 적용 전)
                    json_success, json_msg = self.mask_data_manager.save_masks(
                        self.pdf_manager.file_path,
                        self.masks
                    )
                    if json_success:
                        print(f"마스킹 데이터 JSON 저장: {json_msg}")
                    
                    # 마스킹 적용 및 저장
                    self.pdf_manager.apply_masks_and_save(self.masks)
                    # 마스킹 내역 엑셀 저장
                    self.export_masks_to_excel()
                    
                    # 로그 기록
                    self.log_manager.log_mask_save(
                        self.pdf_manager.file_path,
                        self.masks
                    )
                    
                    # 성공 메시지
                    QMessageBox.information(
                        self,
                        "저장 완료",
                        f"총 {len(self.masks)}개의 마스킹이 적용되어 저장되었습니다."
                    )
                    
                    # 마스킹 데이터 초기화
                    self.clear_masks()
                    
                    # 완료 파일 목록에 추가
                    if self.pdf_manager.file_path:
                        filename = os.path.basename(self.pdf_manager.file_path)
                        if filename not in self.completed_files:
                            self.completed_files.append(filename)
                    
                    # 진행상황 저장
                    if self.current_folder_path and self.pdf_files:
                        self.progress_manager.save_progress(
                            self.current_folder_path,
                            self.pdf_files,
                            self.completed_files,
                            self.current_pdf_index
                        )
                    
                    # 다음 파일로 이동할지 확인
                    self.move_to_next_pdf_if_available()
                    
                    print("마스킹 저장 완료")
                    
                except Exception as e:
                    QMessageBox.critical(
                        self,
                        "저장 오류",
                        f"저장 중 오류가 발생했습니다.\n\n{str(e)}"
                    )
                    print(f"저장 실패: {str(e)}")
        
        # (2) 마스크가 없는 경우
        else:
            reply = QMessageBox.question(
                self,
                "다음으로 이동",
                "마스킹 작업 없이 다음으로 넘어가시겠습니까?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.Yes
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                self.move_to_next_pdf_if_available()

    def move_to_next_pdf_if_available(self) -> None:
        """폴더의 다음 PDF 파일로 이동 (있는 경우)"""
        # 폴더에서 여러 PDF를 열었고, 다음 파일이 있는 경우
        if self.current_pdf_index >= 0 and self.current_pdf_index < len(self.pdf_files) - 1:
            next_index = self.current_pdf_index + 1
            
            reply = QMessageBox.question(
                self,
                "다음 파일",
                f"다음 PDF 파일을 열겠습니까?\n\n{os.path.basename(self.pdf_files[next_index])}",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.Yes
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                self.load_pdf_from_list(next_index)
        else:
            # 마지막 파일인 경우
            if self.current_pdf_index == len(self.pdf_files) - 1:
                # 진행상황 삭제
                self.progress_manager.clear_progress()
                
                QMessageBox.information(
                    self,
                    "완료",
                    "폴더의 모든 PDF 파일 작업이 완료되었습니다."
                )

    def on_mask_created(self, page_index: int, rect: fitz.Rect) -> None:
        """마스킹 영역이 생성되었을 때 호출되는 슬롯"""
        # 새로운 MaskEntry 생성
        mask_entry = MaskEntry(page_index=page_index, rect=rect, note="")
        self.masks.append(mask_entry)
        
        print(f"Mask created on page {page_index + 1}: {rect}")
        
        # 테이블에 행 추가
        row = self.mask_list.rowCount()
        self.mask_list.insertRow(row)
        
        # 페이지 컬럼 (읽기 전용)
        page_item = QTableWidgetItem(str(page_index + 1))
        page_item.setFlags(page_item.flags() & ~Qt.ItemFlag.ItemIsEditable)  # 편집 불가
        self.mask_list.setItem(row, 0, page_item)
        
        # 메모 컬럼 (편집 가능)
        note_item = QTableWidgetItem("")
        self.mask_list.setItem(row, 1, note_item)
        
        # UserRole에 mask 인덱스 저장 (row == mask index)
        page_item.setData(Qt.ItemDataRole.UserRole, row)

    def on_mask_item_changed(self, item: QTableWidgetItem) -> None:
        """마스킹 리스트 아이템이 변경되었을 때 호출되는 슬롯"""
        row = item.row()
        col = item.column()
        
        # 메모 컬럼(1번)만 처리
        if col == 1:
            # row와 masks 인덱스가 1:1 대응
            if 0 <= row < len(self.masks):
                new_note = item.text()
                self.masks[row].note = new_note
                print(f"Mask [{row}] note updated: '{new_note}'")

    def closeEvent(self, event) -> None:
        """윈도우 종료 이벤트"""
        # 로그 기록
        self.log_manager.log_app_end()
        
        # PDF 문서 닫기
        self.pdf_manager.close()
        event.accept()


def main() -> None:
    """메인 함수"""
    app = QApplication(sys.argv)
    app.setApplicationName("PDF Mask")
    
    # 임시 로그 관리자 (라이선스 인증용)
    temp_log = LogManager()
    
    # 라이선스 확인
    license_manager = LicenseManager()
    
    if not license_manager.is_licensed():
        # 라이선스가 없으면 시리얼 번호 입력 다이얼로그 표시
        temp_log.log_license_check(False, "No license file found")
        serial_dialog = SerialInputDialog()
        result = serial_dialog.exec()
        
        if result != QDialog.DialogCode.Accepted:
            # 사용자가 취소하면 프로그램 종료
            temp_log.log_license_check(False, "User cancelled license activation")
            QMessageBox.warning(
                None,
                "종료",
                "라이선스 인증이 필요합니다.\n프로그램을 종료합니다."
            )
            sys.exit(0)
        else:
            temp_log.log_license_check(True, "License activated successfully")
    else:
        temp_log.log_license_check(True, "License verified")
    
    # 메인 윈도우 표시
    window = MainWindow()
    window.show()
    
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
